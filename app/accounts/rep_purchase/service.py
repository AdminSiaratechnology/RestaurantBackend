import io
from datetime import date, datetime, timedelta, timezone
from typing import Optional, List, Tuple

from sqlalchemy import select, func
from sqlalchemy.orm import joinedload, selectinload
from sqlalchemy.ext.asyncio import AsyncSession
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.accounts.branch.model import Branch
from app.accounts.client.model import Client
from app.accounts.purchase.model import PurchaseEntry, PurchaseEntryItem
from app.accounts.vendor.model import Vendor


def resolve_date_range(
    from_date: date | None = None,
    to_date: date | None = None,
    time_range: str | None = None,
) -> tuple[date, date]:
    today = date.today()
    if time_range == "today":
        return today, today
    elif time_range in ("7d", "last_7_days"):
        return today - timedelta(days=6), today
    elif time_range in ("month", "this_month"):
        return today.replace(day=1), today
    elif from_date is not None and to_date is not None:
        if from_date > to_date:
            raise ValueError("from_date must be less than or equal to to_date")
        return from_date, to_date
    elif from_date is not None:
        return from_date, today
    elif to_date is not None:
        return today.replace(day=1), to_date
    else:
        return today.replace(day=1), today


def build_purchase_excel_workbook(
    title: str,
    scope_name: str,
    from_date: date,
    to_date: date,
    purchases: list[PurchaseEntry],
    supplier_name: str | None = None,
) -> io.BytesIO:
    wb = openpyxl.Workbook()

    # --- STYLES DEFINITION ---
    font_title = Font(name="Segoe UI", size=14, bold=True, color="FFFFFF")
    font_subtitle = Font(name="Segoe UI", size=9, italic=True, color="CBD5E1")
    font_kpi_label = Font(name="Segoe UI", size=8, bold=True, color="475569")
    font_kpi_val = Font(name="Segoe UI", size=12, bold=True, color="0F172A")
    font_header = Font(name="Segoe UI", size=9, bold=True, color="FFFFFF")
    font_data = Font(name="Segoe UI", size=9, color="0F172A")
    font_total = Font(name="Segoe UI", size=9, bold=True, color="0F172A")

    fill_title = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    fill_header = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    fill_kpi = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    fill_kpi_highlight = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
    fill_total = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
    fill_zebra = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")

    thin_border_side = Side(style="thin", color="E2E8F0")
    border_cell = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    border_total = Border(top=Side(style="thin", color="D97706"), bottom=Side(style="double", color="D97706"))

    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")

    num_fmt_curr = "₹#,##0.00"
    num_fmt_qty = "#,##0.00"

    # =========================================================================
    # SHEET 1: Purchase Summary
    # =========================================================================
    ws1 = wb.active
    ws1.title = "Purchase Summary"
    ws1.views.sheetView[0].showGridLines = True

    # 1. Title Block
    ws1.merge_cells("A1:S1")
    title_cell = ws1["A1"]
    title_cell.value = f"📊  {title.upper()}"
    title_cell.font = font_title
    title_cell.fill = fill_title
    title_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws1.row_dimensions[1].height = 34

    # 2. Subtitle / Metadata
    ws1.merge_cells("A2:S2")
    sub_cell = ws1["A2"]
    now_str = datetime.now().strftime("%d-%m-%Y %H:%M:%S")
    supp_info = f" | Supplier: {supplier_name}" if supplier_name else ""
    sub_cell.value = f"Scope: {scope_name} | Period: {from_date.strftime('%d-%m-%Y')} to {to_date.strftime('%d-%m-%Y')}{supp_info} | Generated: {now_str}"
    sub_cell.font = font_subtitle
    sub_cell.fill = fill_title
    sub_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws1.row_dimensions[2].height = 20

    # 3. KPI Summary Cards (Row 4 & 5)
    total_entries = len(purchases)
    total_subtotal = sum(float(p.subtotal or 0.0) for p in purchases)
    total_discount = sum(float(p.discount_amount or 0.0) for p in purchases)
    total_tax = sum(float(p.tax_amount or 0.0) for p in purchases)
    total_amount = sum(float(p.grand_total or 0.0) for p in purchases)

    kpis = [
        ("TOTAL ENTRIES", str(total_entries), "A4:C4", "A5:C5", fill_kpi),
        ("TOTAL SUBTOTAL", f"₹{total_subtotal:,.2f}", "D4:F4", "D5:F5", fill_kpi),
        ("TOTAL DISCOUNT", f"₹{total_discount:,.2f}", "G4:I4", "G5:I5", fill_kpi),
        ("TOTAL TAX", f"₹{total_tax:,.2f}", "J4:L4", "J5:L5", fill_kpi),
        ("TOTAL PURCHASE AMOUNT", f"₹{total_amount:,.2f}", "M4:P4", "M5:P5", fill_kpi_highlight),
    ]

    for label, val, top_range, bot_range, kpi_fill in kpis:
        ws1.merge_cells(top_range)
        top_c = ws1[top_range.split(":")[0]]
        top_c.value = label
        top_c.font = font_kpi_label
        top_c.fill = kpi_fill
        top_c.alignment = align_center

        ws1.merge_cells(bot_range)
        bot_c = ws1[bot_range.split(":")[0]]
        bot_c.value = val
        bot_c.font = font_kpi_val
        bot_c.fill = kpi_fill
        bot_c.alignment = align_center

    ws1.row_dimensions[4].height = 18
    ws1.row_dimensions[5].height = 24

    # 4. Table Headers (Row 7)
    headers_ws1 = [
        ("Sr. No.", align_center, 8),
        ("Branch ID", align_center, 12),
        ("Branch Name", align_left, 22),
        ("Purchase ID", align_center, 14),
        ("Invoice Number", align_left, 18),
        ("Invoice Date", align_center, 14),
        ("Supplier ID", align_center, 12),
        ("Supplier Name", align_left, 24),
        ("Supplier Inv No", align_left, 18),
        ("Supplier Inv Date", align_center, 16),
        ("Delivery Date", align_center, 14),
        ("Reference No", align_left, 16),
        ("Payment Terms", align_left, 16),
        ("Due Date", align_center, 14),
        ("Subtotal (₹)", align_right, 16),
        ("Discount (₹)", align_right, 14),
        ("Tax (₹)", align_right, 14),
        ("Total Amount (₹)", align_right, 18),
        ("Notes", align_left, 26),
    ]

    ws1.row_dimensions[7].height = 26
    for col_idx, (h_title, h_align, _) in enumerate(headers_ws1, start=1):
        cell = ws1.cell(row=7, column=col_idx, value=h_title)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = h_align
        cell.border = border_cell

    # 5. Populate Data Rows
    current_row = 8
    for idx, p in enumerate(purchases, start=1):
        ws1.row_dimensions[current_row].height = 20
        row_fill = fill_zebra if idx % 2 == 0 else None

        branch_name = p.branch.name if p.branch else f"Branch #{p.branch_id}"
        supp_name = (
            p.supplier.vendor_name or p.supplier.name
            if p.supplier
            else f"Supplier #{p.supplier_id}"
        )

        inv_d = p.invoice_date.strftime("%d-%m-%Y") if p.invoice_date else "—"
        supp_inv_d = p.supplier_invoice_date.strftime("%d-%m-%Y") if p.supplier_invoice_date else "—"
        deliv_d = p.delivery_date.strftime("%d-%m-%Y") if p.delivery_date else "—"
        due_d = p.due_date.strftime("%d-%m-%Y") if p.due_date else "—"

        subtotal_val = float(p.subtotal or 0.0)
        discount_val = float(p.discount_amount or 0.0)
        tax_val = float(p.tax_amount or 0.0)
        total_val = float(p.grand_total or 0.0)

        row_values = [
            (idx, align_center, None),
            (p.branch_id, align_center, None),
            (branch_name, align_left, None),
            (p.id, align_center, None),
            (p.invoice_number or f"INV-{p.id}", align_left, None),
            (inv_d, align_center, None),
            (p.supplier_id, align_center, None),
            (supp_name, align_left, None),
            (p.supplier_invoice_number or "—", align_left, None),
            (supp_inv_d, align_center, None),
            (deliv_d, align_center, None),
            (p.reference_number or "—", align_left, None),
            (p.payment_terms or "—", align_left, None),
            (due_d, align_center, None),
            (subtotal_val, align_right, num_fmt_curr),
            (discount_val, align_right, num_fmt_curr),
            (tax_val, align_right, num_fmt_curr),
            (total_val, align_right, num_fmt_curr),
            (p.notes or "", align_left, None),
        ]

        for col_idx, (val, cell_align, c_fmt) in enumerate(row_values, start=1):
            cell = ws1.cell(row=current_row, column=col_idx, value=val)
            cell.font = font_data
            cell.alignment = cell_align
            cell.border = border_cell
            if row_fill:
                cell.fill = row_fill
            if c_fmt:
                cell.number_format = c_fmt

        current_row += 1

    # 6. Total Summary Row
    if purchases:
        ws1.row_dimensions[current_row].height = 22
        ws1.cell(row=current_row, column=1, value="TOTAL")
        ws1.cell(row=current_row, column=1).font = font_total
        ws1.cell(row=current_row, column=1).alignment = align_center

        for col_idx in range(1, 20):
            c = ws1.cell(row=current_row, column=col_idx)
            c.fill = fill_total
            c.border = border_total
            c.font = font_total

        ws1.cell(row=current_row, column=15, value=total_subtotal).number_format = num_fmt_curr
        ws1.cell(row=current_row, column=16, value=total_discount).number_format = num_fmt_curr
        ws1.cell(row=current_row, column=17, value=total_tax).number_format = num_fmt_curr
        ws1.cell(row=current_row, column=18, value=total_amount).number_format = num_fmt_curr
    else:
        ws1.row_dimensions[current_row].height = 24
        ws1.merge_cells(f"A{current_row}:S{current_row}")
        empty_c = ws1[f"A{current_row}"]
        empty_c.value = "No purchase entries found for the selected filters."
        empty_c.alignment = align_center
        empty_c.font = font_data

    for col_idx, (_, _, width) in enumerate(headers_ws1, start=1):
        col_letter = get_column_letter(col_idx)
        ws1.column_dimensions[col_letter].width = max(width, 10)

    ws1.freeze_panes = "A8"

    # =========================================================================
    # SHEET 2: Purchase Items Detail
    # =========================================================================
    ws2 = wb.create_sheet(title="Purchase Items")
    ws2.views.sheetView[0].showGridLines = True

    # 1. Title Block
    ws2.merge_cells("A1:S1")
    title2_cell = ws2["A1"]
    title2_cell.value = f"📦  PURCHASE ITEMS DETAIL"
    title2_cell.font = font_title
    title2_cell.fill = fill_title
    title2_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws2.row_dimensions[1].height = 34

    # 2. Subtitle
    ws2.merge_cells("A2:S2")
    sub2_cell = ws2["A2"]
    sub2_cell.value = f"Scope: {scope_name} | Period: {from_date.strftime('%d-%m-%Y')} to {to_date.strftime('%d-%m-%Y')}{supp_info} | Generated: {now_str}"
    sub2_cell.font = font_subtitle
    sub2_cell.fill = fill_title
    sub2_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws2.row_dimensions[2].height = 20

    # 3. Table Headers (Row 4)
    headers_ws2 = [
        ("Sr. No.", align_center, 8),
        ("Branch ID", align_center, 12),
        ("Branch Name", align_left, 22),
        ("Purchase ID", align_center, 14),
        ("Invoice Number", align_left, 18),
        ("Invoice Date", align_center, 14),
        ("Supplier Name", align_left, 24),
        ("Item Name", align_left, 26),
        ("Inventory Item ID", align_center, 16),
        ("Category", align_left, 16),
        ("Godown ID", align_center, 12),
        ("Base Unit", align_center, 12),
        ("Display Unit", align_center, 14),
        ("Conversion Factor", align_right, 16),
        ("Quantity", align_right, 14),
        ("Rate (₹)", align_right, 14),
        ("Discount %", align_right, 14),
        ("Tax %", align_right, 12),
        ("Item Total (₹)", align_right, 18),
    ]

    ws2.row_dimensions[4].height = 26
    for col_idx, (h_title, h_align, _) in enumerate(headers_ws2, start=1):
        cell = ws2.cell(row=4, column=col_idx, value=h_title)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = h_align
        cell.border = border_cell

    # 4. Populate Items
    current_row_items = 5
    item_counter = 1
    total_item_qty = 0.0
    total_item_amount = 0.0

    for p in purchases:
        b_name = p.branch.name if p.branch else f"Branch #{p.branch_id}"
        s_name = (
            p.supplier.vendor_name or p.supplier.name
            if p.supplier
            else f"Supplier #{p.supplier_id}"
        )
        inv_no = p.invoice_number or f"INV-{p.id}"
        inv_dt = p.invoice_date.strftime("%d-%m-%Y") if p.invoice_date else "—"

        if p.items:
            for item in p.items:
                ws2.row_dimensions[current_row_items].height = 20
                r_fill = fill_zebra if item_counter % 2 == 0 else None

                qty = float(item.quantity or 0.0)
                rate = float(item.rate or 0.0)
                disc_pct = float(item.discount_percent or 0.0)
                tax_pct = float(item.tax_percent or 0.0)
                amt = float(item.amount or (qty * rate))
                conv_factor = float(item.conversion_factor or 1.0)

                total_item_qty += qty
                total_item_amount += amt

                item_row_values = [
                    (item_counter, align_center, None),
                    (p.branch_id, align_center, None),
                    (b_name, align_left, None),
                    (p.id, align_center, None),
                    (inv_no, align_left, None),
                    (inv_dt, align_center, None),
                    (s_name, align_left, None),
                    (item.item_name or "Unnamed Item", align_left, None),
                    (item.inventory_item_id or "—", align_center, None),
                    (item.row_category or "—", align_left, None),
                    (item.godown_id or "—", align_center, None),
                    (item.unit or "—", align_center, None),
                    (item.display_unit or "—", align_center, None),
                    (conv_factor, align_right, num_fmt_qty),
                    (qty, align_right, num_fmt_qty),
                    (rate, align_right, num_fmt_curr),
                    (disc_pct, align_right, num_fmt_qty),
                    (tax_pct, align_right, num_fmt_qty),
                    (amt, align_right, num_fmt_curr),
                ]

                for col_idx, (val, cell_align, c_fmt) in enumerate(item_row_values, start=1):
                    cell = ws2.cell(row=current_row_items, column=col_idx, value=val)
                    cell.font = font_data
                    cell.alignment = cell_align
                    cell.border = border_cell
                    if r_fill:
                        cell.fill = r_fill
                    if c_fmt:
                        cell.number_format = c_fmt

                item_counter += 1
                current_row_items += 1

    # Total row on items sheet
    if item_counter > 1:
        ws2.row_dimensions[current_row_items].height = 22
        ws2.cell(row=current_row_items, column=1, value="TOTAL")
        ws2.cell(row=current_row_items, column=1).font = font_total
        ws2.cell(row=current_row_items, column=1).alignment = align_center

        for col_idx in range(1, 20):
            c = ws2.cell(row=current_row_items, column=col_idx)
            c.fill = fill_total
            c.border = border_total
            c.font = font_total

        ws2.cell(row=current_row_items, column=15, value=total_item_qty).number_format = num_fmt_qty
        ws2.cell(row=current_row_items, column=19, value=total_item_amount).number_format = num_fmt_curr
    else:
        ws2.row_dimensions[current_row_items].height = 24
        ws2.merge_cells(f"A{current_row_items}:S{current_row_items}")
        empty_c = ws2[f"A{current_row_items}"]
        empty_c.value = "No purchase items recorded for the selected period."
        empty_c.alignment = align_center
        empty_c.font = font_data

    for col_idx, (_, _, width) in enumerate(headers_ws2, start=1):
        col_letter = get_column_letter(col_idx)
        ws2.column_dimensions[col_letter].width = max(width, 10)

    ws2.freeze_panes = "A5"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


class PurchaseReportService:

    @staticmethod
    async def get_branch_report(
        db: AsyncSession,
        branch_id: int,
        supplier_id: int | None = None,
        from_date: date | None = None,
        to_date: date | None = None,
        time_range: str | None = None,
    ):
        # --------------------------------------------------
        # GET BRANCH
        # --------------------------------------------------
        branch_result = await db.execute(
            select(Branch).where(
                Branch.id == branch_id
            )
        )
        branch = branch_result.scalar_one_or_none()

        if not branch:
            raise ValueError("Branch not found")

        today = date.today()
        seven_days_ago = today - timedelta(days=6)
        month_start = today.replace(day=1)

        f_date, t_date = resolve_date_range(from_date, to_date, time_range)

        # --------------------------------------------------
        # BASE PURCHASE QUERY
        # --------------------------------------------------
        # Fetch records covering at least the month & 7d window for complete KPIs, or custom window
        query_start = min(month_start, seven_days_ago, f_date)
        query_end = max(today, t_date)

        purchase_conditions = [
            PurchaseEntry.branch_id == branch_id,
            PurchaseEntry.invoice_date >= query_start,
            PurchaseEntry.invoice_date <= query_end,
        ]

        if supplier_id is not None:
            purchase_conditions.append(PurchaseEntry.supplier_id == supplier_id)

        purchases_result = await db.execute(
            select(PurchaseEntry)
            .where(*purchase_conditions)
            .order_by(PurchaseEntry.invoice_date.asc())
        )

        purchases = purchases_result.scalars().all()

        # --------------------------------------------------
        # KPI
        # --------------------------------------------------
        today_purchase = 0.0
        last_7_days_purchase = 0.0
        current_month_purchase = 0.0

        for purchase in purchases:
            amount = float(purchase.grand_total or 0)
            purchase_date = purchase.invoice_date

            if purchase_date == today:
                today_purchase += amount

            if seven_days_ago <= purchase_date <= today:
                last_7_days_purchase += amount

            if month_start <= purchase_date <= today:
                current_month_purchase += amount

        # --------------------------------------------------
        # CHARTS - 7 DAYS, MONTH, TODAY
        # --------------------------------------------------

        # 1. 7 Days Chart
        chart_7d = []
        for i in range(7):
            current_date = seven_days_ago + timedelta(days=i)
            amount = 0.0
            for purchase in purchases:
                if purchase.invoice_date == current_date:
                    amount += float(purchase.grand_total or 0)

            if current_date == today:
                label = "Today"
            elif current_date == today - timedelta(days=1):
                label = "Yesterday"
            else:
                days_ago = (today - current_date).days
                label = f"{days_ago} Days Ago"

            chart_7d.append(
                {
                    "date": str(current_date),
                    "label": label,
                    "amount": round(amount, 2),
                }
            )

        # 2. This Month Chart (Weekly Breakdown)
        chart_month = []
        weeks_def = [
            ("Week 1", 1, 7),
            ("Week 2", 8, 14),
            ("Week 3", 15, 21),
            ("Week 4", 22, 28),
            ("Week 5", 29, 31),
        ]
        for w_label, start_d, end_d in weeks_def:
            w_amount = 0.0
            for purchase in purchases:
                p_date = purchase.invoice_date
                if p_date and p_date.month == today.month and p_date.year == today.year:
                    if start_d <= p_date.day <= end_d:
                        w_amount += float(purchase.grand_total or 0)

            chart_month.append(
                {
                    "label": w_label,
                    "amount": round(w_amount, 2),
                }
            )

        # 3. Today Chart (Time Interval Breakdown)
        chart_today = []
        time_slots = [
            ("9 AM", 0, 9),
            ("12 PM", 9, 12),
            ("3 PM", 12, 15),
            ("6 PM", 15, 18),
            ("9 PM", 18, 21),
            ("11 PM", 21, 24),
        ]
        today_purchases = [p for p in purchases if p.invoice_date == today]
        for slot_label, h_start, h_end in time_slots:
            slot_amount = 0.0
            for p in today_purchases:
                if p.created_at:
                    hour = p.created_at.hour
                    if h_start <= hour < h_end:
                        slot_amount += float(p.grand_total or 0)

            chart_today.append(
                {
                    "label": slot_label,
                    "amount": round(slot_amount, 2),
                }
            )

        today_slot_sum = sum(s["amount"] for s in chart_today)
        if today_purchase > 0 and today_slot_sum == 0:
            chart_today[3]["amount"] = round(today_purchase, 2)

        # --------------------------------------------------
        # TOP PURCHASING ITEMS
        # --------------------------------------------------
        item_conditions = [
            PurchaseEntry.branch_id == branch_id,
            PurchaseEntry.invoice_date >= month_start,
            PurchaseEntry.invoice_date <= today,
        ]

        if supplier_id is not None:
            item_conditions.append(PurchaseEntry.supplier_id == supplier_id)

        item_result = await db.execute(
            select(
                PurchaseEntryItem.inventory_item_id,
                PurchaseEntryItem.item_name,
                func.sum(
                    PurchaseEntryItem.quantity
                ).label("total_quantity"),
                func.sum(
                    PurchaseEntryItem.amount
                ).label("total_amount"),
            )
            .join(
                PurchaseEntry,
                PurchaseEntry.id
                == PurchaseEntryItem.purchase_entry_id,
            )
            .where(*item_conditions)
            .group_by(
                PurchaseEntryItem.inventory_item_id,
                PurchaseEntryItem.item_name,
            )
            .order_by(
                func.sum(
                    PurchaseEntryItem.amount
                ).desc()
            )
            .limit(10)
        )

        items = item_result.all()

        total_item_amount = sum(
            float(item.total_amount or 0)
            for item in items
        )

        top_items = []

        for index, item in enumerate(items, start=1):
            amount = float(item.total_amount or 0)
            percentage = (
                (amount / total_item_amount) * 100
                if total_item_amount > 0
                else 0.0
            )

            top_items.append(
                {
                    "rank": index,
                    "inventory_item_id": item.inventory_item_id,
                    "item_name": item.item_name,
                    "total_quantity": round(
                        float(item.total_quantity or 0),
                        2,
                    ),
                    "total_amount": round(
                        amount,
                        2,
                    ),
                    "percentage_of_total": round(
                        percentage,
                        2,
                    ),
                }
            )

        # --------------------------------------------------
        # RESPONSE
        # --------------------------------------------------
        return {
            "branch_id": branch.id,
            "branch_name": branch.name,

            "kpis": {
                "today_purchase": round(
                    today_purchase,
                    2,
                ),
                "last_7_days_purchase": round(
                    last_7_days_purchase,
                    2,
                ),
                "current_month_purchase": round(
                    current_month_purchase,
                    2,
                ),
                "total_purchase_entries": len(
                    purchases
                ),
            },

            "chart": {
                "period": "last_7_days",
                "data": chart_7d,
            },

            "charts": {
                "7d": chart_7d,
                "month": chart_month,
                "today": chart_today,
            },

            "top_purchasing_items": top_items,
        }

    @staticmethod
    async def export_branch_report(
        db: AsyncSession,
        branch_id: int,
        from_date: date | None = None,
        to_date: date | None = None,
        time_range: str | None = None,
        supplier_id: int | None = None,
    ) -> Tuple[io.BytesIO, str]:
        # Validate branch
        branch_result = await db.execute(
            select(Branch).where(Branch.id == branch_id)
        )
        branch = branch_result.scalar_one_or_none()
        if not branch:
            raise ValueError(f"Branch with ID {branch_id} not found")

        # Resolve dates
        f_date, t_date = resolve_date_range(from_date, to_date, time_range)

        # Build query
        conditions = [
            PurchaseEntry.branch_id == branch_id,
            PurchaseEntry.invoice_date >= f_date,
            PurchaseEntry.invoice_date <= t_date,
        ]

        supplier_name = None
        if supplier_id is not None:
            conditions.append(PurchaseEntry.supplier_id == supplier_id)
            vend_res = await db.execute(select(Vendor).where(Vendor.id == supplier_id))
            vend = vend_res.scalar_one_or_none()
            if vend:
                supplier_name = vend.vendor_name or vend.name

        purchases_result = await db.execute(
            select(PurchaseEntry)
            .options(
                joinedload(PurchaseEntry.branch),
                joinedload(PurchaseEntry.supplier),
                selectinload(PurchaseEntry.items),
            )
            .where(*conditions)
            .order_by(
                PurchaseEntry.invoice_date.desc(),
                PurchaseEntry.id.desc(),
            )
        )
        purchases = purchases_result.scalars().all()

        scope_name = f"Branch: {branch.name} (ID: {branch.id})"
        title = f"Purchase Report - {branch.name}"
        filename = f"Purchase_Report_Branch_{branch.id}_{f_date.strftime('%Y%m%d')}_{t_date.strftime('%Y%m%d')}.xlsx"

        excel_buf = build_purchase_excel_workbook(
            title=title,
            scope_name=scope_name,
            from_date=f_date,
            to_date=t_date,
            purchases=purchases,
            supplier_name=supplier_name,
        )

        return excel_buf, filename

    @staticmethod
    async def export_client_report(
        db: AsyncSession,
        client_id: int,
        from_date: date | None = None,
        to_date: date | None = None,
        time_range: str | None = None,
        supplier_id: int | None = None,
        branch_id: int | None = None,
    ) -> Tuple[io.BytesIO, str]:
        # Validate client
        client_res = await db.execute(
            select(Client).where(Client.id == client_id)
        )
        client = client_res.scalar_one_or_none()
        if not client:
            raise ValueError(f"Client with ID {client_id} not found")

        # Get client branches
        branch_query = select(Branch).where(Branch.client_id == client_id)
        if branch_id is not None:
            branch_query = branch_query.where(Branch.id == branch_id)

        branch_res = await db.execute(branch_query.order_by(Branch.id.asc()))
        branches = branch_res.scalars().all()

        if not branches:
            raise ValueError(f"No branches found for client ID {client_id}")

        branch_ids = [b.id for b in branches]

        # Resolve dates
        f_date, t_date = resolve_date_range(from_date, to_date, time_range)

        # Build query
        conditions = [
            PurchaseEntry.branch_id.in_(branch_ids),
            PurchaseEntry.invoice_date >= f_date,
            PurchaseEntry.invoice_date <= t_date,
        ]

        supplier_name = None
        if supplier_id is not None:
            conditions.append(PurchaseEntry.supplier_id == supplier_id)
            vend_res = await db.execute(select(Vendor).where(Vendor.id == supplier_id))
            vend = vend_res.scalar_one_or_none()
            if vend:
                supplier_name = vend.vendor_name or vend.name

        purchases_result = await db.execute(
            select(PurchaseEntry)
            .options(
                joinedload(PurchaseEntry.branch),
                joinedload(PurchaseEntry.supplier),
                selectinload(PurchaseEntry.items),
            )
            .where(*conditions)
            .order_by(
                PurchaseEntry.branch_id.asc(),
                PurchaseEntry.invoice_date.desc(),
                PurchaseEntry.id.desc(),
            )
        )
        purchases = purchases_result.scalars().all()

        scope_name = f"Client: {client.name} ({len(branches)} Branches)"
        if branch_id and len(branches) == 1:
            scope_name = f"Client: {client.name} - Branch: {branches[0].name}"

        title = f"Client Purchase Report - {client.name}"
        filename = f"Purchase_Report_Client_{client.id}_{f_date.strftime('%Y%m%d')}_{t_date.strftime('%Y%m%d')}.xlsx"

        excel_buf = build_purchase_excel_workbook(
            title=title,
            scope_name=scope_name,
            from_date=f_date,
            to_date=t_date,
            purchases=purchases,
            supplier_name=supplier_name,
        )

        return excel_buf, filename