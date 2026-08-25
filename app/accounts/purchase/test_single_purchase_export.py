# app/accounts/purchase/test_single_purchase_export.py

import os
import sys
from datetime import date, datetime
from unittest.mock import MagicMock
import openpyxl

from app.accounts.purchase.model import PurchaseEntry, PurchaseEntryItem
from app.accounts.purchase.service import generate_single_purchase_excel, sanitize_filename
from app.accounts.vendor.model import Vendor


def test_sanitize_filename():
    assert sanitize_filename("INV-0014") == "INV-0014"
    assert sanitize_filename("INV/0014/2026") == "INV_0014_2026"
    assert sanitize_filename("INV #0014 *?") == "INV__0014"
    assert sanitize_filename("") == "invoice"


def test_generate_single_purchase_excel():
    # Mock Vendor
    mock_vendor = Vendor()
    mock_vendor.id = 1
    mock_vendor.name = "Ramesh Suppliers"
    mock_vendor.vendor_name = "Ramesh"
    mock_vendor.phone = "9876543211"

    # Mock Purchase Entry
    purchase = PurchaseEntry()
    purchase.id = 14
    purchase.invoice_number = "INV-0014"
    purchase.invoice_date = date(2026, 8, 22)
    purchase.supplier_invoice_number = "SUP-INV-998"
    purchase.supplier_invoice_date = date(2026, 8, 20)
    purchase.delivery_date = date(2026, 8, 22)
    purchase.reference_number = "PO-2026-88"
    purchase.payment_terms = "Net 30"
    purchase.due_date = date(2026, 9, 21)
    purchase.notes = "Regular grocery purchase"
    purchase.subtotal = 4200.00
    purchase.discount_amount = 0.00
    purchase.tax_amount = 0.00
    purchase.grand_total = 4200.00
    purchase.supplier = mock_vendor

    # Mock Items
    item1 = PurchaseEntryItem()
    item1.id = 1
    item1.item_name = "Wheat Flour"
    item1.row_category = "Groceries"
    item1.quantity = 20.0
    item1.unit = "kg"
    item1.display_unit = "kg"
    item1.conversion_factor = 1.0
    item1.rate = 45.0
    item1.discount_percent = 0.0
    item1.tax_percent = 0.0
    item1.amount = 900.0
    item1.godown = None
    item1.inventory_item = None
    item1.vendor_name = "Ramesh"
    item1.vendor_phone = "9876543211"

    item2 = PurchaseEntryItem()
    item2.id = 2
    item2.item_name = "Rice"
    item2.row_category = "Grains"
    item2.quantity = 25.0
    item2.unit = "kg"
    item2.display_unit = "kg"
    item2.conversion_factor = 1.0
    item2.rate = 60.0
    item2.discount_percent = 0.0
    item2.tax_percent = 0.0
    item2.amount = 1500.0
    item2.godown = None
    item2.inventory_item = None
    item2.vendor_name = "Ramesh"
    item2.vendor_phone = "9876543211"

    item3 = PurchaseEntryItem()
    item3.id = 3
    item3.item_name = "Cooking Oil"
    item3.row_category = "Oils"
    item3.quantity = 10.0
    item3.unit = "L"
    item3.display_unit = "L"
    item3.conversion_factor = 1.0
    item3.rate = 180.0
    item3.discount_percent = 0.0
    item3.tax_percent = 0.0
    item3.amount = 1800.0
    item3.godown = None
    item3.inventory_item = None
    item3.vendor_name = "Ramesh"
    item3.vendor_phone = "9876543211"

    purchase.items = [item1, item2, item3]

    # Generate Excel
    file_path, file_name = generate_single_purchase_excel(purchase)

    assert os.path.exists(file_path), f"File {file_path} does not exist"
    assert file_name == "purchase_invoice_INV-0014.xlsx"

    # Open with openpyxl and verify sheets and structure
    wb = openpyxl.load_workbook(file_path)
    assert "Invoice" in wb.sheetnames, "Missing 'Invoice' sheet"
    assert "Invoice_Items" in wb.sheetnames, "Missing 'Invoice_Items' sheet"
    assert len(wb.sheetnames) == 2, f"Expected 2 sheets, found {len(wb.sheetnames)}"

    # Check Sheet 1: Invoice
    ws1 = wb["Invoice"]
    assert ws1.cell(row=1, column=1).value == "Field"
    assert ws1.cell(row=1, column=2).value == "Value"

    invoice_data = {
        ws1.cell(row=r, column=1).value: ws1.cell(row=r, column=2).value
        for r in range(2, ws1.max_row + 1)
    }

    assert invoice_data["Invoice Number"] == "INV-0014"
    assert invoice_data["Supplier Name"] == "Ramesh"
    assert invoice_data["Invoice Date"] == "22/08/2026"
    assert invoice_data["Supplier Invoice Number"] == "SUP-INV-998"
    assert invoice_data["Grand Total"] == 4200.0
    assert invoice_data["Notes"] == "Regular grocery purchase"

    # Check Sheet 2: Invoice_Items
    ws2 = wb["Invoice_Items"]
    expected_headers = [
        "Item",
        "Category",
        "Quantity",
        "Unit",
        "Display Unit",
        "Conversion Factor",
        "Rate",
        "Discount %",
        "Tax %",
        "Amount",
        "Godown",
        "Vendor Name",
        "Vendor Phone",
    ]
    actual_headers = [ws2.cell(row=1, column=c).value for c in range(1, len(expected_headers) + 1)]
    assert actual_headers == expected_headers, f"Headers mismatch: {actual_headers}"

    # Verify rows
    assert ws2.cell(row=2, column=1).value == "Wheat Flour"
    assert ws2.cell(row=2, column=3).value == 20.0
    assert ws2.cell(row=2, column=4).value == "kg"
    assert ws2.cell(row=2, column=7).value == 45.0
    assert ws2.cell(row=2, column=10).value == 900.0

    assert ws2.cell(row=3, column=1).value == "Rice"
    assert ws2.cell(row=3, column=3).value == 25.0
    assert ws2.cell(row=3, column=7).value == 60.0
    assert ws2.cell(row=3, column=10).value == 1500.0

    assert ws2.cell(row=4, column=1).value == "Cooking Oil"
    assert ws2.cell(row=4, column=3).value == 10.0
    assert ws2.cell(row=4, column=7).value == 180.0
    assert ws2.cell(row=4, column=10).value == 1800.0

    # Total Row (Row 5)
    assert ws2.cell(row=5, column=1).value == "Total"
    assert ws2.cell(row=5, column=3).value == 55.0  # 20 + 25 + 10
    assert ws2.cell(row=5, column=10).value == 4200.0  # 900 + 1500 + 1800

    print("All assertions in test_generate_single_purchase_excel passed successfully!")


if __name__ == "__main__":
    test_sanitize_filename()
    test_generate_single_purchase_excel()
    print("ALL TESTS PASSED!")
