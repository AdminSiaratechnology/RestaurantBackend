import os
import re
from datetime import date, datetime, timezone
from typing import List, Optional, Tuple

from fastapi import HTTPException, status
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from sqlalchemy import or_, select
from sqlalchemy.dialects.postgresql import insert as pg_insert
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.accounts.branch.model import Branch

from app.accounts.inventory.model import (
    Godown,
    InventoryItem,
)

from app.accounts.purchase.model import (
    BranchPurchaseInvoiceCounter,
    PurchaseEntry,
    PurchaseEntryItem,
)

from app.accounts.purchase.schema import (
    PurchaseCreate,
    PurchaseUpdate,
)

from app.accounts.vendor.model import Vendor


# ============================================================
# INVOICE NUMBER FORMATTING
# ============================================================

def format_invoice_number(
    number: int,
) -> str:
    return f"INV-{number:04d}"


# ============================================================
# VALIDATE BRANCH
# ============================================================

async def validate_branch(
    db: AsyncSession,
    branch_id: int,
) -> None:

    if branch_id <= 0:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Invalid branch_id",
        )

    result = await db.execute(
        select(Branch.id).where(
            Branch.id == branch_id
        )
    )

    if result.scalar_one_or_none() is None:

        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Branch not found",
        )


# ============================================================
# VALIDATE SUPPLIER
# ============================================================

async def validate_supplier(
    db: AsyncSession,
    supplier_id: int,
    branch_id: int,
) -> Vendor:

    result = await db.execute(
        select(Vendor).where(
            Vendor.id == supplier_id,
            or_(
                Vendor.branch_id == branch_id,
                Vendor.branch_id.is_(None),
            ),
        )
    )

    vendor = result.scalar_one_or_none()

    if vendor is None:
        # Check if vendor exists in database
        vendor = await db.get(Vendor, supplier_id)

    if vendor is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=(
                "Supplier/Vendor not found"
            ),
        )

    return vendor



# ============================================================
# VALIDATE GODOWN
# ============================================================

async def validate_godown(
    db: AsyncSession,
    branch_id: int,
    godown_id: int,
) -> Godown:

    godown = await db.get(
        Godown,
        godown_id,
    )

    if not godown:

        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Godown not found",
        )

    if godown.branch_id != branch_id:

        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=(
                "Godown does not belong "
                "to selected branch"
            ),
        )

    return godown


# ============================================================
# NEXT INVOICE PREVIEW
# ============================================================

async def get_next_invoice_preview(
    db: AsyncSession,
    branch_id: int,
) -> dict:

    await validate_branch(
        db,
        branch_id,
    )

    result = await db.execute(
        select(
            BranchPurchaseInvoiceCounter.last_invoice_number
        ).where(
            BranchPurchaseInvoiceCounter.branch_id
            == branch_id
        )
    )

    last_number = (
        result.scalar_one_or_none()
    )

    next_number = (
        1
        if last_number is None
        else last_number + 1
    )

    return {
        "branch_id": branch_id,
        "invoice_number":
            format_invoice_number(next_number),
        "invoice_date": date.today(),
    }


# ============================================================
# RESERVE INVOICE NUMBER
# ============================================================

async def reserve_next_invoice_number(
    db: AsyncSession,
    branch_id: int,
) -> int:

    stmt = (
        pg_insert(
            BranchPurchaseInvoiceCounter
        )
        .values(
            branch_id=branch_id,
            last_invoice_number=1,
        )
        .on_conflict_do_update(
            index_elements=[
                BranchPurchaseInvoiceCounter.branch_id
            ],
            set_={
                "last_invoice_number":
                    BranchPurchaseInvoiceCounter
                    .last_invoice_number + 1,

                "updated_at":
                    datetime.now(timezone.utc),
            },
        )
        .returning(
            BranchPurchaseInvoiceCounter
            .last_invoice_number
        )
    )

    result = await db.execute(stmt)

    return result.scalar_one()


async def generate_next_purchase_invoice_number(
    db: AsyncSession,
    branch_id: int,
) -> int:

    return await reserve_next_invoice_number(
        db,
        branch_id,
    )


# ============================================================
# CALCULATE ITEM AMOUNT
# ============================================================

def calculate_item_amount(
    quantity: float,
    rate: float,
    discount_percent: float,
    tax_percent: float,
) -> float:

    gross_amount = (
        quantity * rate
    )

    discount_amount = (
        gross_amount
        * discount_percent
        / 100
    )

    taxable_amount = (
        gross_amount
        - discount_amount
    )

    tax_amount = (
        taxable_amount
        * tax_percent
        / 100
    )

    return round(
        taxable_amount + tax_amount,
        2,
    )


# ============================================================
# CREATE PURCHASE
# ============================================================

async def create_purchase_entry(
    db: AsyncSession,
    payload: PurchaseCreate,
) -> PurchaseEntry:

    # ========================================================
    # 1. VALIDATE BRANCH
    # ========================================================

    await validate_branch(
        db,
        payload.branch_id,
    )

    # ========================================================
    # 2. VALIDATE SUPPLIER
    # ========================================================

    await validate_supplier(
        db,
        payload.supplier_id,
        payload.branch_id,
    )

    # ========================================================
    # 3. VALIDATE TOTAL
    # ========================================================

    calculated_total = round(
        payload.subtotal
        + payload.tax_amount
        - payload.discount_amount,
        2,
    )

    if abs(
        calculated_total - payload.grand_total
    ) > 0.01:

        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail={
                "message": "Grand total does not match calculated total",
                "expected_grand_total": calculated_total,
                "received_grand_total": payload.grand_total,
            },
        )

    # ========================================================
    # 4. GENERATE INVOICE
    # ========================================================

    next_number = await reserve_next_invoice_number(
        db,
        payload.branch_id,
    )

    invoice_number = format_invoice_number(
        next_number
    )

    invoice_date = date.today()

    # ========================================================
    # 5. CREATE PURCHASE HEADER
    # ========================================================

    purchase = PurchaseEntry(
        branch_id=payload.branch_id,
        supplier_id=payload.supplier_id,

        invoice_number=invoice_number,
        invoice_date=invoice_date,

        supplier_invoice_number=(
            payload.supplier_invoice_number
        ),

        supplier_invoice_date=(
            payload.supplier_invoice_date
        ),

        delivery_date=payload.delivery_date,

        reference_number=payload.reference_number,

        payment_terms=payload.payment_terms,

        due_date=payload.due_date,

        notes=payload.notes,

        subtotal=payload.subtotal,

        tax_amount=payload.tax_amount,

        discount_amount=payload.discount_amount,

        grand_total=payload.grand_total,
    )

    db.add(purchase)

    await db.flush()

    # ========================================================
    # 6. ADD PURCHASE ITEMS
    # ========================================================

    for item_data in payload.items:

        inventory_item = None

        # ----------------------------------------------------
        # If inventory item selected
        # ----------------------------------------------------

        if item_data.inventory_item_id is not None:

            result = await db.execute(
                select(InventoryItem)
                .where(
                    InventoryItem.id
                    == item_data.inventory_item_id
                )
                .with_for_update()
            )

            inventory_item = (
                result.scalar_one_or_none()
            )

            if inventory_item is None:

                raise HTTPException(
                    status_code=status.HTTP_404_NOT_FOUND,
                    detail=(
                        f"Inventory item "
                        f"{item_data.inventory_item_id} "
                        f"not found"
                    ),
                )

            # Branch validation
            if (
                inventory_item.branch_id
                != payload.branch_id
            ):

                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail=(
                        "Inventory item does not "
                        "belong to selected branch"
                    ),
                )

        # ----------------------------------------------------
        # Validate godown
        # ----------------------------------------------------

        if item_data.godown_id is not None:

            await validate_godown(
                db,
                payload.branch_id,
                item_data.godown_id,
            )

        elif inventory_item is not None:

            item_data.godown_id = (
                inventory_item.godown_id
            )

        # ----------------------------------------------------
        # Snapshot values from inventory
        # ----------------------------------------------------

        item_name = item_data.item_name

        row_category = (
            item_data.row_category
            or (
                inventory_item.row_category
                if inventory_item
                else "other"
            )
        )

        unit = (
            item_data.unit
            or (
                inventory_item.unit
                if inventory_item
                else "piece"
            )
        )

        display_unit = (
            item_data.display_unit
            or (
                inventory_item.display_unit
                if inventory_item
                else unit
            )
        )

        conversion_factor = (
            item_data.conversion_factor
            or (
                inventory_item.conversion_factor
                if inventory_item
                else 1
            )
        )

        reorder_level = (
            item_data.reorder_level
            if item_data.reorder_level is not None
            else (
                inventory_item.reorder_level
                if inventory_item
                else 0
            )
        )

        vendor_name = (
            item_data.vendor_name
            or (
                inventory_item.vendor_name
                if inventory_item
                else None
            )
        )

        vendor_phone = (
            item_data.vendor_phone
            or (
                inventory_item.vendor_phone
                if inventory_item
                else None
            )
        )

        # ----------------------------------------------------
        # Calculate amount
        # ----------------------------------------------------

        gross_amount = (
            item_data.quantity
            * item_data.rate
        )

        discount_amount = (
            gross_amount
            * item_data.discount_percent
            / 100
        )

        taxable_amount = (
            gross_amount
            - discount_amount
        )

        tax_amount = (
            taxable_amount
            * item_data.tax_percent
            / 100
        )

        calculated_item_amount = round(
            taxable_amount + tax_amount,
            2,
        )

        # If frontend sends amount, validate/use it
        if item_data.amount is not None:

            if abs(
                item_data.amount
                - calculated_item_amount
            ) > 0.01:

                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail={
                        "message": (
                            f"Amount mismatch for "
                            f"item '{item_name}'"
                        ),
                        "expected_amount": (
                            calculated_item_amount
                        ),
                        "received_amount": (
                            item_data.amount
                        ),
                    },
                )

            item_amount = item_data.amount

        else:
            item_amount = calculated_item_amount

        # ====================================================
        # CREATE PURCHASE ITEM
        # ====================================================

        purchase_item = PurchaseEntryItem(

            purchase_entry_id=purchase.id,

            inventory_item_id=(
                item_data.inventory_item_id
            ),

            godown_id=(
                item_data.godown_id
            ),

            item_name=item_name,

            row_category=row_category,

            unit=unit,

            display_unit=display_unit,

            conversion_factor=conversion_factor,

            quantity=item_data.quantity,

            reorder_level=reorder_level,

            rate=item_data.rate,

            vendor_name=vendor_name,

            vendor_phone=vendor_phone,

            discount_percent=(
                item_data.discount_percent
            ),

            tax_percent=(
                item_data.tax_percent
            ),

            amount=item_amount,
        )

        db.add(purchase_item)

        # ====================================================
        # UPDATE INVENTORY
        # ====================================================

        if inventory_item is not None:

            # Purchase quantity is entered in display unit.
            # Convert to base unit before storing stock.
            base_quantity = (
                item_data.quantity
                * conversion_factor
            )

            inventory_item.stock_qty = (
                (inventory_item.stock_qty or 0)
                + base_quantity
            )

            inventory_item.cost_per_unit = (
                item_data.rate
                / conversion_factor
                if conversion_factor > 0
                else item_data.rate
            )

            inventory_item.reorder_level = (
                reorder_level
            )

            inventory_item.vendor_name = (
                vendor_name
            )

            inventory_item.vendor_phone = (
                vendor_phone
            )

            inventory_item.last_restocked = (
                datetime.now(timezone.utc)
            )

            inventory_item.status = (
                "in_stock"
                if inventory_item.stock_qty > 0
                else "out_of_stock"
            )

    await db.flush()

    stmt = (
        select(PurchaseEntry)
        .options(
            selectinload(PurchaseEntry.items).selectinload(PurchaseEntryItem.inventory_item),
            selectinload(PurchaseEntry.items).selectinload(PurchaseEntryItem.godown),
        )
        .where(PurchaseEntry.id == purchase.id)
    )
    result = await db.execute(stmt)
    return result.scalar_one()


# ============================================================
# CANONICAL ALIAS
# ============================================================

create_purchase = create_purchase_entry


# ============================================================
# GET PURCHASES
# ============================================================

async def get_purchases(
    db: AsyncSession,
    branch_id: Optional[int] = None,
    supplier_id: Optional[int] = None,
    skip: int = 0,
    limit: int = 50,
) -> List[PurchaseEntry]:

    stmt = (
        select(PurchaseEntry)
        .options(
            selectinload(PurchaseEntry.items).selectinload(PurchaseEntryItem.inventory_item),
            selectinload(PurchaseEntry.items).selectinload(PurchaseEntryItem.godown),
        )
        .order_by(
            PurchaseEntry.id.desc()
        )
    )

    if branch_id is not None:
        stmt = stmt.where(
            PurchaseEntry.branch_id == branch_id
        )

    if supplier_id is not None:
        stmt = stmt.where(
            PurchaseEntry.supplier_id == supplier_id
        )

    stmt = stmt.offset(skip).limit(limit)
    result = await db.execute(stmt)
    return list(result.scalars().all())


get_purchase_entries = get_purchases


# ============================================================
# GET PURCHASE BY ID
# ============================================================

async def get_purchase_by_id(
    db: AsyncSession,
    purchase_id: int,
) -> Optional[PurchaseEntry]:

    stmt = (
        select(PurchaseEntry)
        .options(
            selectinload(PurchaseEntry.items).selectinload(PurchaseEntryItem.inventory_item),
            selectinload(PurchaseEntry.items).selectinload(PurchaseEntryItem.godown),
        )
        .where(
            PurchaseEntry.id == purchase_id
        )
    )

    result = await db.execute(stmt)
    return result.scalar_one_or_none()


get_purchase_entry_by_id = get_purchase_by_id


# ============================================================
# UPDATE PURCHASE
# ============================================================

async def update_purchase(
    db: AsyncSession,
    purchase_id: int,
    payload: PurchaseUpdate,
) -> PurchaseEntry:

    purchase = await get_purchase_by_id(
        db,
        purchase_id,
    )

    if not purchase:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Purchase entry not found",
        )

    # --------------------------------------------------------
    # Supplier
    # --------------------------------------------------------

    if payload.supplier_id is not None:
        await validate_supplier(
            db,
            payload.supplier_id,
            purchase.branch_id,
        )
        purchase.supplier_id = payload.supplier_id

    # --------------------------------------------------------
    # Header fields
    # --------------------------------------------------------

    if payload.supplier_invoice_number is not None:
        purchase.supplier_invoice_number = payload.supplier_invoice_number

    if payload.supplier_invoice_date is not None:
        purchase.supplier_invoice_date = payload.supplier_invoice_date

    if payload.delivery_date is not None:
        purchase.delivery_date = payload.delivery_date

    if payload.reference_number is not None:
        purchase.reference_number = payload.reference_number

    if payload.payment_terms is not None:
        purchase.payment_terms = payload.payment_terms

    if payload.due_date is not None:
        purchase.due_date = payload.due_date

    if payload.notes is not None:
        purchase.notes = payload.notes

    if payload.subtotal is not None:
        purchase.subtotal = payload.subtotal

    if payload.tax_amount is not None:
        purchase.tax_amount = payload.tax_amount

    if payload.discount_amount is not None:
        purchase.discount_amount = payload.discount_amount

    if payload.grand_total is not None:
        purchase.grand_total = payload.grand_total

    purchase.updated_at = datetime.now(timezone.utc)
    await db.flush()

    stmt = (
        select(PurchaseEntry)
        .options(
            selectinload(PurchaseEntry.items).selectinload(PurchaseEntryItem.inventory_item),
            selectinload(PurchaseEntry.items).selectinload(PurchaseEntryItem.godown),
        )
        .where(PurchaseEntry.id == purchase.id)
    )
    result = await db.execute(stmt)
    return result.scalar_one()


# ============================================================
# DELETE PURCHASE
# ============================================================

async def delete_purchase(
    db: AsyncSession,
    purchase_id: int,
) -> bool:

    purchase = await get_purchase_by_id(
        db,
        purchase_id,
    )

    if not purchase:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Purchase entry not found",
        )

    await db.delete(purchase)
    await db.flush()
    return True


# ============================================================
# PURCHASE ITEM LOOKUP
# ============================================================

async def get_purchase_items_lookup(
    db: AsyncSession,
    branch_id: int,
    godown_id: Optional[int] = None,
    search: Optional[str] = None,
) -> List[dict]:

    await validate_branch(db, branch_id)

    query = (
        select(InventoryItem)
        .where(
            InventoryItem.branch_id == branch_id,
        )
        .order_by(
            InventoryItem.name.asc()
        )
    )

    if godown_id is not None:
        godown = await db.get(Godown, godown_id)
        if godown and godown.branch_id == branch_id:
            query = query.where(InventoryItem.godown_id == godown_id)
        else:
            # Godown ID is not in this branch (e.g. initial fallback ID from frontend)
            # Find godowns for this branch
            branch_godowns_res = await db.execute(
                select(Godown.id).where(Godown.branch_id == branch_id)
            )
            b_gids = branch_godowns_res.scalars().all()
            if b_gids:
                query = query.where(
                    or_(
                        InventoryItem.godown_id.in_(b_gids),
                        InventoryItem.godown_id.is_(None),
                    )
                )

    if search:
        query = query.where(
            InventoryItem.name.ilike(
                f"%{search.strip()}%"
            )
        )

    result = await db.execute(query)
    items = result.scalars().all()

    response = []

    for item in items:
        factor = item.conversion_factor or 1

        response.append({
            "id": item.id,
            "inventory_item_id": item.id,
            "name": item.name,
            "item_name": item.name,
            "branch_id": item.branch_id,
            "godown_id": item.godown_id,
            "row_category": item.row_category or "other",
            "unit": item.unit,
            "display_unit": item.display_unit or item.unit,
            "base_unit": item.unit,
            "conversion_factor": factor,
            "current_stock": (item.stock_qty or 0) / factor,
            "stock_qty": item.stock_qty or 0,
            "reorder_level": item.reorder_level or 0,
            "cost_per_unit": item.cost_per_unit or 0,
            "vendor_name": item.vendor_name,
            "vendor_phone": item.vendor_phone,
            "status": item.status,
            "last_restocked": item.last_restocked,
        })

    return response


# ============================================================
# SINGLE PURCHASE EXPORT EXCEL
# ============================================================

def sanitize_filename(name: str) -> str:
    sanitized = re.sub(r"[^a-zA-Z0-9_-]", "_", str(name))
    return sanitized.strip("_") or "invoice"


async def get_purchase_for_export(
    db: AsyncSession,
    purchase_id: int,
) -> Optional[PurchaseEntry]:
    stmt = (
        select(PurchaseEntry)
        .options(
            selectinload(PurchaseEntry.supplier),
            selectinload(PurchaseEntry.branch).selectinload(Branch.client),
            selectinload(PurchaseEntry.items).selectinload(PurchaseEntryItem.inventory_item),
            selectinload(PurchaseEntry.items).selectinload(PurchaseEntryItem.godown),
        )
        .where(PurchaseEntry.id == purchase_id)
    )
    result = await db.execute(stmt)
    return result.scalar_one_or_none()


def generate_single_purchase_excel(
    purchase: PurchaseEntry,
) -> Tuple[str, str]:
    wb = openpyxl.Workbook()

    # Typography & Colors
    FONT_HEADER = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
    FONT_BOLD = Font(name="Segoe UI", size=9, bold=True, color="0F172A")
    FONT_DATA = Font(name="Segoe UI", size=9, color="0F172A")
    FONT_TOTAL = Font(name="Segoe UI", size=10, bold=True, color="0F172A")

    FILL_HEADER = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    FILL_TOTAL = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
    FILL_ZEBRA = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")

    BORDER_THIN = Side(style="thin", color="E2E8F0")
    BORDER_CELL = Border(
        left=BORDER_THIN,
        right=BORDER_THIN,
        top=BORDER_THIN,
        bottom=BORDER_THIN,
    )
    BORDER_TOTAL = Border(
        top=Side(style="thin", color="D97706"),
        bottom=Side(style="double", color="D97706"),
        left=BORDER_THIN,
        right=BORDER_THIN,
    )

    ALIGN_LEFT = Alignment(horizontal="left", vertical="center")
    ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")
    ALIGN_CENTER = Alignment(horizontal="center", vertical="center")

    NUM_FMT_CURRENCY = "₹#,##0.00"
    NUM_FMT_QTY = "#,##0.00"
    NUM_FMT_PERCENT = "0.00%"

    def format_date_str(d):
        if not d:
            return "—"
        if isinstance(d, (datetime, date)):
            return d.strftime("%d/%m/%Y")
        return str(d)

    # Extract supplier name & phone
    vendor_name = "—"
    vendor_phone = "—"
    if purchase.supplier:
        vendor_name = getattr(purchase.supplier, "vendor_name", None) or getattr(purchase.supplier, "name", None) or "—"
        vendor_phone = getattr(purchase.supplier, "phone", None) or getattr(purchase.supplier, "mobile", None) or getattr(purchase.supplier, "contact_number", None) or "—"

    # ========================================================
    # SHEET 1: Invoice
    # ========================================================
    ws1 = wb.active
    ws1.title = "Invoice"
    ws1.views.sheetView[0].showGridLines = True

    # Header Row
    ws1.append(["Field", "Value"])
    ws1.row_dimensions[1].height = 24
    for col_idx in range(1, 3):
        cell = ws1.cell(row=1, column=col_idx)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_LEFT
        cell.border = BORDER_CELL

    summary_fields = [
        ("Invoice Number", purchase.invoice_number, "text"),
        ("Supplier Name", vendor_name, "text"),
        ("Invoice Date", format_date_str(purchase.invoice_date), "text"),
        ("Supplier Invoice Number", purchase.supplier_invoice_number or "—", "text"),
        ("Supplier Invoice Date", format_date_str(purchase.supplier_invoice_date), "text"),
        ("Delivery Date", format_date_str(purchase.delivery_date), "text"),
        ("Reference Number", purchase.reference_number or "—", "text"),
        ("Payment Terms", purchase.payment_terms or "—", "text"),
        ("Due Date", format_date_str(purchase.due_date), "text"),
        ("Notes", purchase.notes or "—", "text"),
        ("Subtotal", float(purchase.subtotal or 0.0), "currency"),
        ("Discount Amount", float(purchase.discount_amount or 0.0), "currency"),
        ("Tax Amount", float(purchase.tax_amount or 0.0), "currency"),
        ("Grand Total", float(purchase.grand_total or 0.0), "currency"),
    ]

    for row_idx, (field, val, val_type) in enumerate(summary_fields, start=2):
        ws1.row_dimensions[row_idx].height = 20
        c1 = ws1.cell(row=row_idx, column=1, value=field)
        c1.font = FONT_BOLD
        c1.alignment = ALIGN_LEFT
        c1.border = BORDER_CELL

        c2 = ws1.cell(row=row_idx, column=2, value=val)
        c2.font = FONT_DATA
        c2.border = BORDER_CELL

        if val_type == "currency":
            c2.number_format = NUM_FMT_CURRENCY
            c2.alignment = ALIGN_RIGHT
            if field == "Grand Total":
                c1.fill = FILL_TOTAL
                c2.fill = FILL_TOTAL
                c2.font = FONT_TOTAL
                c1.border = BORDER_TOTAL
                c2.border = BORDER_TOTAL
        else:
            c2.alignment = ALIGN_LEFT

    ws1.column_dimensions["A"].width = 28
    ws1.column_dimensions["B"].width = 36

    # ========================================================
    # SHEET 2: Invoice_Items
    # ========================================================
    ws2 = wb.create_sheet(title="Invoice_Items")
    ws2.views.sheetView[0].showGridLines = True
    ws2.freeze_panes = "A2"

    item_headers = [
        "Item",
        "Category",
        "Quantity",
        "Unit",
        "Display Unit",
        "Conversion Factor",
        "Rate",
        "Discount %",
        "Tax %",
        "Amount",
        "Godown",
        "Vendor Name",
        "Vendor Phone",
    ]

    ws2.append(item_headers)
    ws2.row_dimensions[1].height = 24
    for col_idx in range(1, len(item_headers) + 1):
        cell = ws2.cell(row=1, column=col_idx)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = (
            ALIGN_CENTER
            if col_idx in (4, 5, 8, 9)
            else (ALIGN_RIGHT if col_idx in (3, 6, 7, 10) else ALIGN_LEFT)
        )
        cell.border = BORDER_CELL

    items = list(purchase.items or [])
    total_qty = 0.0
    total_amount = 0.0

    for r_idx, item in enumerate(items, start=2):
        ws2.row_dimensions[r_idx].height = 20
        is_zebra = (r_idx % 2 == 1)

        item_qty = float(item.quantity or 0.0)
        item_rate = float(item.rate or 0.0)
        item_amt = float(item.amount or (item_qty * item_rate))
        total_qty += item_qty
        total_amount += item_amt

        category = item.row_category or (
            item.inventory_item.row_category if item.inventory_item else "other"
        )
        unit = item.unit or "—"
        display_unit = item.display_unit or unit
        conv_factor = float(item.conversion_factor or 1.0)
        discount_pct = float(item.discount_percent or 0.0)
        tax_pct = float(item.tax_percent or 0.0)
        godown_name = (
            item.godown.name
            if item.godown
            else (
                getattr(item.godown, "godown_name", None)
                if getattr(item, "godown", None)
                else "—"
            )
        ) if getattr(item, "godown", None) else "—"
        v_name = item.vendor_name or vendor_name
        v_phone = item.vendor_phone or vendor_phone

        row_vals = [
            (item.item_name or "—", ALIGN_LEFT, None),
            (category, ALIGN_LEFT, None),
            (item_qty, ALIGN_RIGHT, NUM_FMT_QTY),
            (unit, ALIGN_CENTER, None),
            (display_unit, ALIGN_CENTER, None),
            (conv_factor, ALIGN_RIGHT, "#,##0.##"),
            (item_rate, ALIGN_RIGHT, NUM_FMT_CURRENCY),
            (discount_pct / 100.0, ALIGN_RIGHT, NUM_FMT_PERCENT),
            (tax_pct / 100.0, ALIGN_RIGHT, NUM_FMT_PERCENT),
            (item_amt, ALIGN_RIGHT, NUM_FMT_CURRENCY),
            (godown_name, ALIGN_LEFT, None),
            (v_name, ALIGN_LEFT, None),
            (v_phone, ALIGN_LEFT, None),
        ]

        for c_idx, (v, align, num_fmt) in enumerate(row_vals, start=1):
            cell = ws2.cell(row=r_idx, column=c_idx, value=v)
            cell.font = FONT_DATA
            cell.alignment = align
            cell.border = BORDER_CELL
            if is_zebra:
                cell.fill = FILL_ZEBRA
            if num_fmt:
                cell.number_format = num_fmt

    # Totals Row in Invoice_Items
    if items:
        tot_row = len(items) + 2
        ws2.row_dimensions[tot_row].height = 22
        for col_idx in range(1, len(item_headers) + 1):
            cell = ws2.cell(row=tot_row, column=col_idx)
            cell.font = FONT_TOTAL
            cell.border = BORDER_TOTAL
            cell.fill = FILL_TOTAL
            if col_idx == 1:
                cell.value = "Total"
                cell.alignment = ALIGN_LEFT
            elif col_idx == 3:
                cell.value = total_qty
                cell.alignment = ALIGN_RIGHT
                cell.number_format = NUM_FMT_QTY
            elif col_idx == 10:
                cell.value = total_amount
                cell.alignment = ALIGN_RIGHT
                cell.number_format = NUM_FMT_CURRENCY
            else:
                cell.value = ""

    # Auto-fit column widths for Sheet 2
    for col in ws2.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or "")
            if cell.number_format == NUM_FMT_CURRENCY and isinstance(cell.value, (int, float)):
                val_str = f"₹{cell.value:,.2f}"
            max_len = max(max_len, len(val_str))
        ws2.column_dimensions[col_letter].width = max(max_len + 4, 12)

    # Save to temp directory
    temp_dir = os.path.join(
        os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(__file__)))),
        "temp",
    )
    os.makedirs(temp_dir, exist_ok=True)
    clean_inv = sanitize_filename(purchase.invoice_number or f"INV_{purchase.id}")
    file_name = f"purchase_invoice_{clean_inv}.xlsx"
    file_path = os.path.join(temp_dir, file_name)

    wb.save(file_path)
    return file_path, file_name
