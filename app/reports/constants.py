# app/reports/constants.py

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# =========================================================================
# OPENPYXL STYLES FOR STANDARDIZED EXCEL REPORTS (MATCHING PURCHASE REPORT)
# =========================================================================

FONT_TITLE = Font(name="Segoe UI", size=14, bold=True, color="FFFFFF")
FONT_SUBTITLE = Font(name="Segoe UI", size=9, italic=True, color="CBD5E1")
FONT_KPI_LABEL = Font(name="Segoe UI", size=8, bold=True, color="475569")
FONT_KPI_VAL = Font(name="Segoe UI", size=12, bold=True, color="0F172A")
FONT_HEADER = Font(name="Segoe UI", size=9, bold=True, color="FFFFFF")
FONT_DATA = Font(name="Segoe UI", size=9, color="0F172A")
FONT_TOTAL = Font(name="Segoe UI", size=9, bold=True, color="0F172A")
FONT_EMPTY = Font(name="Segoe UI", size=9, italic=True, color="64748B")

# Fills
FILL_TITLE = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
FILL_HEADER = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
FILL_KPI = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
FILL_KPI_HIGHLIGHT = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
FILL_TOTAL = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
FILL_ZEBRA = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")

# Borders
BORDER_THIN_SIDE = Side(style="thin", color="E2E8F0")
BORDER_CELL = Border(
    left=BORDER_THIN_SIDE,
    right=BORDER_THIN_SIDE,
    top=BORDER_THIN_SIDE,
    bottom=BORDER_THIN_SIDE,
)
BORDER_TOTAL = Border(
    top=Side(style="thin", color="D97706"),
    bottom=Side(style="double", color="D97706"),
)

# Alignments
ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
ALIGN_LEFT = Alignment(horizontal="left", vertical="center")
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")

# Number Formats
NUM_FMT_CURRENCY = "₹#,##0.00"
NUM_FMT_QTY = "#,##0.00"
NUM_FMT_INT = "#,##0"
NUM_FMT_PERCENT = "0.00%"
NUM_FMT_DATE = "DD-MM-YYYY"
NUM_FMT_DATETIME = "DD-MM-YYYY HH:MM:SS"

# Default Item Emojis
DEFAULT_ITEM_EMOJIS = [
    "🌾", "🥛", "🧈", "🥫", "🍅", "🥔", "🧀", "🍗",
    "🥩", "🍚", "🥖", "🥬", "🫒", "🌶️", "🧅", "🧄", "📦", "🏷️"
]
