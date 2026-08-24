# app/reports/export_engine.py

import io
from datetime import date, datetime
from typing import List, Tuple, Optional, Any, Dict
import openpyxl
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

from app.reports.constants import (
    FONT_TITLE,
    FONT_SUBTITLE,
    FONT_KPI_LABEL,
    FONT_KPI_VAL,
    FONT_HEADER,
    FONT_DATA,
    FONT_TOTAL,
    FONT_EMPTY,
    FILL_TITLE,
    FILL_HEADER,
    FILL_KPI,
    FILL_KPI_HIGHLIGHT,
    FILL_TOTAL,
    FILL_ZEBRA,
    BORDER_CELL,
    BORDER_TOTAL,
    ALIGN_CENTER,
    ALIGN_LEFT,
    ALIGN_RIGHT,
    NUM_FMT_CURRENCY,
    NUM_FMT_QTY,
    NUM_FMT_PERCENT,
)


class ExcelReportBuilder:
    """
    Standardized 2-Sheet Excel Report Generator matching the Purchase Report reference design.
    - Sheet 1: Summary with Branded Header, Scope, 3-5 KPI Summary Cards, Summary Data Table, and Double-Border Total Row
    - Sheet 2: Details / Breakdown with Branded Header, Scope, Detailed Items Table, and Double-Border Total Row
    """

    def __init__(
        self,
        report_title: str,
        scope_name: str,
        from_date: date,
        to_date: date,
        filter_subtitle_extra: Optional[str] = None,
    ):
        self.report_title = report_title
        self.scope_name = scope_name
        self.from_date = from_date
        self.to_date = to_date
        self.filter_subtitle_extra = filter_subtitle_extra or ""
        self.wb = openpyxl.Workbook()
        self.generated_at = datetime.now().strftime("%d-%m-%Y %H:%M:%S")

    def _apply_header_block(self, ws, title_text: str, max_col_letter: str = "S"):
        # Row 1: Title
        ws.merge_cells(f"A1:{max_col_letter}1")
        title_cell = ws["A1"]
        title_cell.value = f"📊  {title_text.upper()}"
        title_cell.font = FONT_TITLE
        title_cell.fill = FILL_TITLE
        title_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[1].height = 34

        # Row 2: Scope / Metadata
        ws.merge_cells(f"A2:{max_col_letter}2")
        sub_cell = ws["A2"]
        date_str = f"{self.from_date.strftime('%d-%m-%Y')} to {self.to_date.strftime('%d-%m-%Y')}"
        sub_cell.value = f"Scope: {self.scope_name} | Period: {date_str}{self.filter_subtitle_extra} | Generated: {self.generated_at}"
        sub_cell.font = FONT_SUBTITLE
        sub_cell.fill = FILL_TITLE
        sub_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[2].height = 20

    def add_summary_sheet(
        self,
        sheet_title: str,
        kpis: List[Tuple[str, str, bool]],  # (Label, FormattedValue, is_highlight)
        headers: List[Tuple[str, Alignment, int]],  # (HeaderTitle, Alignment, col_width)
        data_rows: List[List[Tuple[Any, Alignment, Optional[str]]]],  # (Value, Alignment, num_format)
        totals_row: Optional[Dict[int, Tuple[Any, Optional[str]]]] = None,  # col_idx (1-based) -> (val, num_format)
        empty_message: str = "No records found for the selected period.",
    ):
        ws = self.wb.active
        ws.title = sheet_title
        ws.views.sheetView[0].showGridLines = True

        max_col_idx = max(len(headers), 15)
        max_col_letter = get_column_letter(max_col_idx)

        # 1. Header Block
        self._apply_header_block(ws, f"{self.report_title} - {self.scope_name}", max_col_letter)

        # 2. KPI Summary Cards (Row 4 & 5)
        if kpis:
            # Distribute KPI boxes across columns
            # Standard: 3-5 KPI boxes. Each box spans 3-4 columns
            kpi_ranges = [
                ("A4:C4", "A5:C5"),
                ("D4:F4", "D5:F5"),
                ("G4:I4", "G5:I5"),
                ("J4:L4", "J5:L5"),
                ("M4:P4", "M5:P5"),
                ("Q4:S4", "Q5:S5"),
            ]
            for idx, (label, val_str, is_high) in enumerate(kpis):
                if idx >= len(kpi_ranges):
                    break
                top_range, bot_range = kpi_ranges[idx]
                kpi_fill = FILL_KPI_HIGHLIGHT if is_high else FILL_KPI

                ws.merge_cells(top_range)
                top_c = ws[top_range.split(":")[0]]
                top_c.value = label.upper()
                top_c.font = FONT_KPI_LABEL
                top_c.fill = kpi_fill
                top_c.alignment = ALIGN_CENTER

                ws.merge_cells(bot_range)
                bot_c = ws[bot_range.split(":")[0]]
                bot_c.value = val_str
                bot_c.font = FONT_KPI_VAL
                bot_c.fill = kpi_fill
                bot_c.alignment = ALIGN_CENTER

            ws.row_dimensions[4].height = 18
            ws.row_dimensions[5].height = 24

        # 3. Table Headers (Row 7)
        table_start_row = 7
        ws.row_dimensions[table_start_row].height = 26
        for col_idx, (h_title, h_align, _) in enumerate(headers, start=1):
            cell = ws.cell(row=table_start_row, column=col_idx, value=h_title)
            cell.font = FONT_HEADER
            cell.fill = FILL_HEADER
            cell.alignment = h_align
            cell.border = BORDER_CELL

        # 4. Data Rows
        current_row = table_start_row + 1
        for row_idx, row_vals in enumerate(data_rows, start=1):
            ws.row_dimensions[current_row].height = 20
            row_fill = FILL_ZEBRA if row_idx % 2 == 0 else None

            for col_idx, (val, cell_align, c_fmt) in enumerate(row_vals, start=1):
                cell = ws.cell(row=current_row, column=col_idx, value=val)
                cell.font = FONT_DATA
                cell.alignment = cell_align
                cell.border = BORDER_CELL
                if row_fill:
                    cell.fill = row_fill
                if c_fmt:
                    cell.number_format = c_fmt

            current_row += 1

        # 5. Totals Row or Empty Row
        if data_rows:
            if totals_row is not None:
                ws.row_dimensions[current_row].height = 22
                ws.cell(row=current_row, column=1, value="TOTAL").alignment = ALIGN_CENTER

                for col_idx in range(1, len(headers) + 1):
                    c = ws.cell(row=current_row, column=col_idx)
                    c.fill = FILL_TOTAL
                    c.border = BORDER_TOTAL
                    c.font = FONT_TOTAL

                for col_idx, (tot_val, tot_fmt) in totals_row.items():
                    c = ws.cell(row=current_row, column=col_idx, value=tot_val)
                    c.font = FONT_TOTAL
                    c.alignment = ALIGN_RIGHT
                    if tot_fmt:
                        c.number_format = tot_fmt
        else:
            ws.row_dimensions[current_row].height = 24
            last_col_ltr = get_column_letter(len(headers))
            ws.merge_cells(f"A{current_row}:{last_col_ltr}{current_row}")
            empty_c = ws[f"A{current_row}"]
            empty_c.value = empty_message
            empty_c.alignment = ALIGN_CENTER
            empty_c.font = FONT_EMPTY

        # Column widths & freeze panes
        for col_idx, (_, _, width) in enumerate(headers, start=1):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = max(width, 10)

        ws.freeze_panes = f"A{table_start_row + 1}"

    def add_details_sheet(
        self,
        sheet_title: str,
        details_header_title: str,
        headers: List[Tuple[str, Alignment, int]],
        data_rows: List[List[Tuple[Any, Alignment, Optional[str]]]],
        totals_row: Optional[Dict[int, Tuple[Any, Optional[str]]]] = None,
        empty_message: str = "No detailed records found for the selected period.",
    ):
        ws = self.wb.create_sheet(title=sheet_title)
        ws.views.sheetView[0].showGridLines = True

        max_col_idx = max(len(headers), 15)
        max_col_letter = get_column_letter(max_col_idx)

        # 1. Header Block
        self._apply_header_block(ws, details_header_title, max_col_letter)

        # 2. Table Headers (Row 4)
        table_start_row = 4
        ws.row_dimensions[table_start_row].height = 26
        for col_idx, (h_title, h_align, _) in enumerate(headers, start=1):
            cell = ws.cell(row=table_start_row, column=col_idx, value=h_title)
            cell.font = FONT_HEADER
            cell.fill = FILL_HEADER
            cell.alignment = h_align
            cell.border = BORDER_CELL

        # 3. Data Rows
        current_row = table_start_row + 1
        for row_idx, row_vals in enumerate(data_rows, start=1):
            ws.row_dimensions[current_row].height = 20
            row_fill = FILL_ZEBRA if row_idx % 2 == 0 else None

            for col_idx, (val, cell_align, c_fmt) in enumerate(row_vals, start=1):
                cell = ws.cell(row=current_row, column=col_idx, value=val)
                cell.font = FONT_DATA
                cell.alignment = cell_align
                cell.border = BORDER_CELL
                if row_fill:
                    cell.fill = row_fill
                if c_fmt:
                    cell.number_format = c_fmt

            current_row += 1

        # 4. Totals Row or Empty Row
        if data_rows:
            if totals_row is not None:
                ws.row_dimensions[current_row].height = 22
                ws.cell(row=current_row, column=1, value="TOTAL").alignment = ALIGN_CENTER

                for col_idx in range(1, len(headers) + 1):
                    c = ws.cell(row=current_row, column=col_idx)
                    c.fill = FILL_TOTAL
                    c.border = BORDER_TOTAL
                    c.font = FONT_TOTAL

                for col_idx, (tot_val, tot_fmt) in totals_row.items():
                    c = ws.cell(row=current_row, column=col_idx, value=tot_val)
                    c.font = FONT_TOTAL
                    c.alignment = ALIGN_RIGHT
                    if tot_fmt:
                        c.number_format = tot_fmt
        else:
            ws.row_dimensions[current_row].height = 24
            last_col_ltr = get_column_letter(len(headers))
            ws.merge_cells(f"A{current_row}:{last_col_ltr}{current_row}")
            empty_c = ws[f"A{current_row}"]
            empty_c.value = empty_message
            empty_c.alignment = ALIGN_CENTER
            empty_c.font = FONT_EMPTY

        # Column widths & freeze panes
        for col_idx, (_, _, width) in enumerate(headers, start=1):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = max(width, 10)

        ws.freeze_panes = f"A{table_start_row + 1}"

    def build(self) -> io.BytesIO:
        output = io.BytesIO()
        self.wb.save(output)
        output.seek(0)
        return output
